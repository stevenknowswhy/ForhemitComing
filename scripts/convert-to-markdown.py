#!/usr/bin/env python3
"""
Forhemit Document Converter
───────────────────────────
Recursively converts a folder of mixed-format documents into clean markdown
files suitable for agent consumption by /forhemit-preflight and /forhemit-proposal.

Supported formats:
  PDF (.pdf)              → pdfplumber text + table extraction
  Word (.docx)            → pandoc CLI
  Excel (.xlsx, .xls)     → openpyxl → markdown tables (one sheet per section)
  PowerPoint (.pptx)      → pandoc CLI
  CSV (.csv)              → Python csv module → markdown table
  Email (.eml)            → Python email module → headers + body
  Email (.msg)            → extract-msg → headers + body
  HTML (.html, .htm)      → html2text
  Text (.txt)             → pass-through with frontmatter
  Markdown (.md)          → pass-through with frontmatter added
  Images (.jpg, .png, .tiff, .bmp) → OCR via pytesseract (optional)

Usage:
  python scripts/convert-to-markdown.py /path/to/folder
  python scripts/convert-to-markdown.py /path/to/folder --output /path/to/output
  python scripts/convert-to-markdown.py /path/to/folder --dry-run

Output:
  <folder>/output/
    document1.md
    document2.md
    spreadsheet.md
    _manifest.json          ← file map: original → converted path + metadata
"""

import argparse
import csv
import hashlib
import io
import json
import os
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Optional

# ── Optional imports (graceful degradation) ──────────────────────────────────

try:
    import pdfplumber

    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import extract_msg

    HAS_EXTRACT_MSG = True
except ImportError:
    HAS_EXTRACT_MSG = False

try:
    import html2text

    HAS_HTML2TEXT = True
except ImportError:
    HAS_HTML2TEXT = False

try:
    import pytesseract
    from PIL import Image

    HAS_OCR = True
except ImportError:
    HAS_OCR = False

# ── Constants ────────────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".pptx",
    ".xlsx",
    ".xls",
    ".csv",
    ".eml",
    ".msg",
    ".html",
    ".htm",
    ".txt",
    ".md",
    ".jpg",
    ".jpeg",
    ".png",
    ".tiff",
    ".tif",
    ".bmp",
}

SKIP_DIRS = {
    "output",
    ".git",
    "node_modules",
    "__pycache__",
    ".pi",
    ".atl",
    ".pi-lens",
    ".worktrees",
}

MAX_FILE_SIZE_MB = 100  # skip files larger than this


# ── Utilities ────────────────────────────────────────────────────────────────


def file_hash(path: Path) -> str:
    """SHA-256 hash of file contents for dedup and change detection."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def safe_filename(original: Path, base_dir: Path) -> str:
    """
    Generate a safe, flat markdown filename from the original path.
    Preserves enough hierarchy to be unique without nested directories.
    """
    rel = original.relative_to(base_dir)
    parts = list(rel.parts)
    # Replace directory separators with double-dash
    name = "--".join(parts)
    # Strip original extension, add .md
    stem = Path(name).stem
    # Sanitize: lowercase, replace spaces/special chars
    safe = stem.lower().replace(" ", "-")
    # Remove characters that aren't alphanumeric, dash, or underscore
    safe = "".join(c for c in safe if c.isalnum() or c in "-_")
    # Truncate to 120 chars to avoid filesystem limits
    if len(safe) > 120:
        safe = safe[:120]
    return f"{safe}.md"


def frontmatter(source_file: str, file_type: str, extra: Optional[dict] = None) -> str:
    """Generate YAML frontmatter for converted files."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    fm = [
        "---",
        f'source_file: "{source_file}"',
        f'file_type: "{file_type}"',
        f'converted_at: "{now}"',
        'converter: "forhemit-convert v1.0.0"',
    ]
    if extra:
        for k, v in extra.items():
            if isinstance(v, str):
                fm.append(f'{k}: "{v}"')
            elif isinstance(v, (int, float, bool)):
                fm.append(f"{k}: {v}")
            elif isinstance(v, list):
                fm.append(f"{k}:")
                for item in v:
                    fm.append(f'  - "{item}"')
    fm.append("---")
    fm.append("")
    return "\n".join(fm)


def write_markdown(output_dir: Path, filename: str, content: str) -> Path:
    """Write a markdown file to the output directory."""
    out_path = output_dir / filename
    # Handle collisions by appending a counter
    counter = 1
    original_stem = out_path.stem
    while out_path.exists():
        out_path = output_dir / f"{original_stem}-{counter}.md"
        counter += 1
    out_path.write_text(content, encoding="utf-8")
    return out_path


# ── Converters ───────────────────────────────────────────────────────────────


def convert_pdf(source: Path, output_dir: Path) -> dict:
    """Convert PDF to markdown using pdfplumber."""
    if not HAS_PDFPLUMBER:
        return {
            "status": "skipped",
            "reason": "pdfplumber not installed. Run: pip install pdfplumber",
        }

    try:
        pages_text = []
        tables_found = 0

        with pdfplumber.open(source) as pdf:
            total_pages = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                page_parts = []

                # Extract text
                text = page.extract_text()
                if text:
                    page_parts.append(text.strip())

                # Extract tables
                tables = page.extract_tables()
                if tables:
                    tables_found += len(tables)
                    for table in tables:
                        # Convert table to markdown
                        md_table = table_to_markdown(table)
                        if md_table:
                            page_parts.append(md_table)

                if page_parts:
                    pages_text.append(
                        f"<!-- Page {i + 1} of {total_pages} -->\n\n"
                        + "\n\n".join(page_parts)
                    )

        if not pages_text:
            return {
                "status": "empty",
                "reason": "No extractable text found in PDF (may be scanned/image-only)",
                "ocr_available": HAS_OCR,
            }

        content = "\n\n---\n\n".join(pages_text)
        md = frontmatter(
            str(source.name),
            "pdf",
            {
                "pages": total_pages,
                "tables_found": tables_found,
            },
        )
        md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "pages": total_pages,
            "tables": tables_found,
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_pdf_ocr(source: Path, output_dir: Path) -> dict:
    """Convert scanned PDF to markdown using OCR (pytesseract)."""
    if not HAS_OCR:
        return {
            "status": "skipped",
            "reason": "OCR not available. Install: brew install tesseract && pip install pytesseract Pillow",
        }

    try:
        from pdf2image import convert_from_path

        images = convert_from_path(str(source), dpi=200)
        pages_text = []

        for i, img in enumerate(images):
            text = pytesseract.image_to_string(img)
            if text.strip():
                pages_text.append(f"<!-- Page {i + 1} (OCR) -->\n\n{text.strip()}")

        if not pages_text:
            return {"status": "empty", "reason": "OCR produced no extractable text"}

        content = "\n\n---\n\n".join(pages_text)
        md = frontmatter(str(source.name), "pdf-ocr", {"pages": len(images)})
        md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "pages": len(images),
            "method": "ocr",
            "size_bytes": out_path.stat().st_size,
        }

    except ImportError:
        return {
            "status": "skipped",
            "reason": "pdf2image not installed. Run: pip install pdf2image",
        }
    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_docx(source: Path, output_dir: Path) -> dict:
    """Convert DOCX to markdown using pandoc."""
    pandoc_path = shutil.which("pandoc")
    if not pandoc_path:
        return {
            "status": "skipped",
            "reason": "pandoc not installed. Run: brew install pandoc",
        }

    try:
        result = subprocess.run(
            [pandoc_path, "-f", "docx", "-t", "markdown", "--wrap=none", str(source)],
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode != 0:
            return {"status": "error", "reason": result.stderr.strip()}

        content = result.stdout
        if not content.strip():
            return {"status": "empty", "reason": "No text content found in DOCX"}

        md = frontmatter(str(source.name), "docx")
        md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except subprocess.TimeoutExpired:
        return {"status": "error", "reason": "Pandoc conversion timed out (60s)"}
    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_pptx(source: Path, output_dir: Path) -> dict:
    """Convert PowerPoint to markdown using pandoc."""
    pandoc_path = shutil.which("pandoc")
    if not pandoc_path:
        return {
            "status": "skipped",
            "reason": "pandoc not installed. Run: brew install pandoc",
        }

    try:
        result = subprocess.run(
            [pandoc_path, "-f", "pptx", "-t", "markdown", "--wrap=none", str(source)],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            return {"status": "error", "reason": result.stderr.strip()}

        content = result.stdout
        if not content.strip():
            return {"status": "empty", "reason": "No text content found in PPTX"}

        md = frontmatter(str(source.name), "pptx")
        md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except subprocess.TimeoutExpired:
        return {"status": "error", "reason": "Pandoc conversion timed out (120s)"}
    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_xlsx(source: Path, output_dir: Path) -> dict:
    """Convert Excel workbook to markdown tables (one section per sheet)."""
    if not HAS_OPENPYXL:
        return {
            "status": "skipped",
            "reason": "openpyxl not installed. Run: pip install openpyxl",
        }

    try:
        wb = openpyxl.load_workbook(str(source), data_only=True, read_only=True)
        sheets_content = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                rows.append([str(cell) if cell is not None else "" for cell in row])

            if not rows:
                continue

            total_rows += len(rows)

            # Build markdown table
            sheet_md = f"## Sheet: {sheet_name}\n\n"
            sheet_md += table_rows_to_markdown(rows)
            sheets_content.append(sheet_md)

        wb.close()

        if not sheets_content:
            return {"status": "empty", "reason": "No data found in any sheet"}

        content = "\n\n".join(sheets_content)
        md = frontmatter(
            str(source.name),
            "xlsx",
            {
                "sheets": wb.sheetnames,
                "total_rows": total_rows,
            },
        )
        md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "sheets": len(sheets_content),
            "rows": total_rows,
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_csv(source: Path, output_dir: Path) -> dict:
    """Convert CSV to markdown table."""
    try:
        # Try UTF-8 first, fall back to latin-1
        try:
            text = source.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = source.read_text(encoding="latin-1")

        reader = csv.reader(io.StringIO(text))
        rows = [row for row in reader if any(cell.strip() for cell in row)]

        if not rows:
            return {"status": "empty", "reason": "CSV file is empty"}

        content = table_rows_to_markdown(rows)
        md = frontmatter(str(source.name), "csv", {"rows": len(rows)})
        md += f"# {source.stem}\n\n{content}"

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "rows": len(rows),
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_eml(source: Path, output_dir: Path) -> dict:
    """Convert .eml email file to markdown."""
    try:
        with open(source, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        headers = {
            "From": str(msg.get("From", "")),
            "To": str(msg.get("To", "")),
            "Cc": str(msg.get("Cc", "")),
            "Subject": str(msg.get("Subject", "")),
            "Date": str(msg.get("Date", "")),
        }

        # Get body
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                ct = part.get_content_type()
                if ct == "text/plain":
                    body = part.get_content()
                    break
                elif ct == "text/html" and not body:
                    raw_html = part.get_content()
                    if HAS_HTML2TEXT:
                        h = html2text.HTML2Text()
                        h.ignore_links = False
                        body = h.handle(raw_html)
                    else:
                        body = raw_html
        else:
            body = msg.get_content() or ""

        md = frontmatter(
            str(source.name),
            "eml",
            {
                "email_from": headers["From"],
                "email_to": headers["To"],
                "email_subject": headers["Subject"],
                "email_date": headers["Date"],
            },
        )
        md += f"# Email: {headers['Subject']}\n\n"
        md += f"**From:** {headers['From']}  \n"
        md += f"**To:** {headers['To']}  \n"
        if headers["Cc"]:
            md += f"**Cc:** {headers['Cc']}  \n"
        md += f"**Date:** {headers['Date']}  \n\n"
        md += "---\n\n"
        md += body.strip()

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_msg(source: Path, output_dir: Path) -> dict:
    """Convert .msg (Outlook) email file to markdown."""
    if not HAS_EXTRACT_MSG:
        return {
            "status": "skipped",
            "reason": "extract-msg not installed. Run: pip install extract-msg",
        }

    try:
        msg = extract_msg.Message(str(source))

        headers = {
            "From": msg.sender or "",
            "To": msg.to or "",
            "Cc": msg.cc or "",
            "Subject": msg.subject or "",
            "Date": msg.date or "",
        }

        body = msg.body or ""

        md = frontmatter(
            str(source.name),
            "msg",
            {
                "email_from": headers["From"],
                "email_to": headers["To"],
                "email_subject": headers["Subject"],
                "email_date": headers["Date"],
            },
        )
        md += f"# Email: {headers['Subject']}\n\n"
        md += f"**From:** {headers['From']}  \n"
        md += f"**To:** {headers['To']}  \n"
        if headers["Cc"]:
            md += f"**Cc:** {headers['Cc']}  \n"
        md += f"**Date:** {headers['Date']}  \n\n"
        md += "---\n\n"
        md += body.strip()

        msg.close()

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_html(source: Path, output_dir: Path) -> dict:
    """Convert HTML to markdown."""
    if not HAS_HTML2TEXT:
        return {
            "status": "skipped",
            "reason": "html2text not installed. Run: pip install html2text",
        }

    try:
        try:
            raw = source.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            raw = source.read_text(encoding="latin-1")

        h = html2text.HTML2Text()
        h.ignore_links = False
        h.ignore_images = False
        h.body_width = 0  # no wrapping
        content = h.handle(raw)

        if not content.strip():
            return {"status": "empty", "reason": "No text content found in HTML"}

        md = frontmatter(str(source.name), "html")
        md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_text(source: Path, output_dir: Path) -> dict:
    """Convert plain text file to markdown."""
    try:
        try:
            content = source.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            content = source.read_text(encoding="latin-1")

        if not content.strip():
            return {"status": "empty", "reason": "File is empty"}

        md = frontmatter(str(source.name), "text")
        md += f"# {source.stem}\n\n```\n{content}\n```"

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_markdown(source: Path, output_dir: Path) -> dict:
    """Copy markdown file with added frontmatter (if not already present)."""
    try:
        content = source.read_text(encoding="utf-8")

        # Add frontmatter only if not already present
        if content.startswith("---"):
            md = content
        else:
            md = frontmatter(str(source.name), "markdown")
            md += content

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


def convert_image_ocr(source: Path, output_dir: Path) -> dict:
    """OCR an image file to markdown."""
    if not HAS_OCR:
        return {
            "status": "skipped",
            "reason": "OCR not available. Install: brew install tesseract && pip install pytesseract Pillow",
        }

    try:
        img = Image.open(str(source))
        text = pytesseract.image_to_string(img)

        if not text.strip():
            return {"status": "empty", "reason": "OCR produced no extractable text"}

        md = frontmatter(str(source.name), "image-ocr", {"method": "tesseract"})
        md += f"# OCR: {source.name}\n\n{text.strip()}"

        filename = safe_filename(source, source.parent)
        out_path = write_markdown(output_dir, filename, md)

        return {
            "status": "converted",
            "output": str(out_path.name),
            "method": "ocr",
            "size_bytes": out_path.stat().st_size,
        }

    except Exception as e:
        return {"status": "error", "reason": str(e)}


# ── Table Formatting ─────────────────────────────────────────────────────────


def table_to_markdown(table: list) -> str:
    """Convert a pdfplumber table (list of lists) to markdown table string."""
    if not table or not table[0]:
        return ""

    # Clean cells
    cleaned = []
    for row in table:
        cleaned.append(
            [
                str(cell).replace("\n", " ").replace("|", "\\|").strip() if cell else ""
                for cell in row
            ]
        )

    return table_rows_to_markdown(cleaned)


def table_rows_to_markdown(rows: list) -> str:
    """Convert a list of row lists to a markdown table."""
    if not rows or not rows[0]:
        return ""

    num_cols = max(len(row) for row in rows)

    # Normalize column count
    normalized = []
    for row in rows:
        padded = row + [""] * (num_cols - len(row))
        normalized.append(padded)

    lines = []
    # Header row
    lines.append("| " + " | ".join(normalized[0]) + " |")
    lines.append("| " + " | ".join(["---"] * num_cols) + " |")
    # Data rows
    for row in normalized[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


# ── Dispatcher ───────────────────────────────────────────────────────────────

CONVERTERS = {
    ".pdf": convert_pdf,
    ".docx": convert_docx,
    ".pptx": convert_pptx,
    ".xlsx": convert_xlsx,
    ".xls": convert_xlsx,
    ".csv": convert_csv,
    ".eml": convert_eml,
    ".msg": convert_msg,
    ".html": convert_html,
    ".htm": convert_html,
    ".txt": convert_text,
    ".md": convert_markdown,
    ".jpg": convert_image_ocr,
    ".jpeg": convert_image_ocr,
    ".png": convert_image_ocr,
    ".tiff": convert_image_ocr,
    ".tif": convert_image_ocr,
    ".bmp": convert_image_ocr,
}


# ── Main ─────────────────────────────────────────────────────────────────────


def scan_folder(folder: Path) -> list[Path]:
    """Recursively find all supported files in the folder."""
    found = []
    for root, dirs, files in os.walk(folder):
        # Skip excluded directories
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
        for fname in sorted(files):
            fpath = Path(root) / fname
            ext = fpath.suffix.lower()
            if ext in SUPPORTED_EXTENSIONS:
                # Skip files that are too large
                size_mb = fpath.stat().st_size / (1024 * 1024)
                if size_mb > MAX_FILE_SIZE_MB:
                    continue
                found.append(fpath)
    return found


def convert_folder(
    folder: Path, output_dir: Optional[Path] = None, dry_run: bool = False
) -> dict:
    """
    Convert all supported files in a folder to markdown.

    Returns a manifest dict with conversion results.
    """
    folder = folder.resolve()
    if output_dir is None:
        output_dir = folder / "output"
    else:
        output_dir = output_dir.resolve()

    # Scan
    files = scan_folder(folder)
    print(f"\n  Found {len(files)} supported file(s) in {folder}\n")

    if not files:
        print("  No supported files found. Nothing to convert.")
        return {
            "source_folder": str(folder),
            "output_folder": str(output_dir),
            "files": [],
        }

    if dry_run:
        print("  DRY RUN — no files will be written\n")

    # Create output directory
    if not dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)

    # Convert
    manifest = {
        "source_folder": str(folder),
        "output_folder": str(output_dir),
        "converted_at": datetime.now(timezone.utc).isoformat(),
        "total_files": len(files),
        "converted": 0,
        "skipped": 0,
        "errors": 0,
        "empty": 0,
        "ocr_available": HAS_OCR,
        "dependencies": {
            "pdfplumber": HAS_PDFPLUMBER,
            "openpyxl": HAS_OPENPYXL,
            "extract_msg": HAS_EXTRACT_MSG,
            "html2text": HAS_HTML2TEXT,
            "ocr": HAS_OCR,
            "pandoc": shutil.which("pandoc") is not None,
        },
        "files": [],
    }

    for fpath in files:
        ext = fpath.suffix.lower()
        rel_path = str(fpath.relative_to(folder))
        converter = CONVERTERS.get(ext)

        entry = {
            "source": rel_path,
            "extension": ext,
            "size_bytes": fpath.stat().st_size,
            "hash": file_hash(fpath),
        }

        if converter is None:
            entry["status"] = "unsupported"
            entry["reason"] = f"No converter for {ext}"
            manifest["skipped"] += 1
            print(f"  ⏭  {rel_path}  (no converter for {ext})")
        elif dry_run:
            entry["status"] = "would_convert"
            print(f"  🔍 {rel_path}  → would convert via {ext}")
        else:
            print(f"  🔄 {rel_path}  ", end="", flush=True)
            try:
                result = converter(fpath, output_dir)
                entry.update(result)
                status = result.get("status", "unknown")
                if status == "converted":
                    manifest["converted"] += 1
                    print(f"→ {result.get('output', '?')}")
                elif status == "empty":
                    manifest["empty"] += 1
                    print(f"  (empty: {result.get('reason', '')})")
                elif status == "skipped":
                    manifest["skipped"] += 1
                    print(f"  (skipped: {result.get('reason', '')})")
                else:
                    manifest["errors"] += 1
                    print(f"  ERROR: {result.get('reason', 'unknown')}")
            except Exception as e:
                entry["status"] = "error"
                entry["reason"] = str(e)
                manifest["errors"] += 1
                print(f"  ERROR: {e}")

        manifest["files"].append(entry)

    # Write manifest
    if not dry_run:
        manifest_path = output_dir / "_manifest.json"
        manifest_path.write_text(
            json.dumps(manifest, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(f"\n  📋 Manifest: {manifest_path}")

    # Summary
    print("\n  ── Summary ──")
    print(f"  Total files:   {manifest['total_files']}")
    print(f"  Converted:     {manifest['converted']}")
    print(f"  Empty:         {manifest['empty']}")
    print(f"  Skipped:       {manifest['skipped']}")
    print(f"  Errors:        {manifest['errors']}")
    print(f"  Output:        {output_dir}")
    print()

    return manifest


def main():
    parser = argparse.ArgumentParser(
        description="Forhemit Document Converter — convert mixed-format folders to markdown"
    )
    parser.add_argument(
        "folder",
        type=str,
        help="Path to the folder containing documents to convert",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default=None,
        help="Output directory for markdown files (default: <folder>/markdown/)",
    )
    parser.add_argument(
        "--dry-run",
        "-n",
        action="store_true",
        help="Show what would be converted without writing files",
    )
    parser.add_argument(
        "--check-deps",
        action="store_true",
        help="Check dependency availability and exit",
    )

    args = parser.parse_args()

    if args.check_deps:
        print("\n  Forhemit Document Converter — Dependency Check\n")
        deps = [
            ("pdfplumber", HAS_PDFPLUMBER, "pip install pdfplumber"),
            ("openpyxl", HAS_OPENPYXL, "pip install openpyxl"),
            ("extract-msg", HAS_EXTRACT_MSG, "pip install extract-msg"),
            ("html2text", HAS_HTML2TEXT, "pip install html2text"),
            ("pandoc", shutil.which("pandoc") is not None, "brew install pandoc"),
            (
                "pytesseract + Pillow",
                HAS_OCR,
                "brew install tesseract && pip install pytesseract Pillow",
            ),
        ]
        for name, available, install in deps:
            icon = "✅" if available else "❌"
            print(f"  {icon}  {name:<25} {'installed' if available else install}")
        print()
        return

    folder = Path(args.folder).resolve()
    if not folder.exists():
        print(f"\n  ERROR: Folder not found: {folder}\n")
        sys.exit(1)
    if not folder.is_dir():
        print(f"\n  ERROR: Not a directory: {folder}\n")
        sys.exit(1)

    output_dir = Path(args.output).resolve() if args.output else None

    convert_folder(folder, output_dir, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
